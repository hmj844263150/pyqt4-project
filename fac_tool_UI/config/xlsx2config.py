import openpyxl
import os

print ('start')
for root, dirs, files in os.walk('./'):
    for fn in files:
        if fn.endswith('.xlsx'):
            wb = openpyxl.load_workbook(fn)
            sh1 = wb.get_sheet_names()[0]
            sheet1 = wb.get_sheet_by_name(sh1)
            with open('config', 'w') as fd:
                pass
            
            #print (sheet1.cell(row=1, column=1).value)
            #sheet1.cell(row=1,column=2).value = 'hehe'
            #wb.save(fn)
            
print ('fin')