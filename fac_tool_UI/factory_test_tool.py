import sys
import os
from factory_test_tool_ui import *
import time
import ConfigParser
import serial
import serial.tools.list_ports 
import threading
import datetime
import Queue
import win32api
import requests
import json
import inspect  
import ctypes
import res_rc
from memory_profiler import profile
from my_widget import *
from io import StringIO

sys.path.append('../')
import RF_test.esp_test as esp_test

sys.path.append('../')
import upload_to_server.upload_to_server as upload_to_server
from upload_to_server import *

from PyQt4 import QtCore, QtGui

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)
    
    




class FactoryToolUI(Ui_MainWindow, QtGui.QMainWindow):
    _SP_SIGN = '$$'
    SIGNAL_PRINT = QtCore.pyqtSignal(QtGui.QTextBrowser, str)
    DUT_NUM = 4
    DUT_PORTS = []
    DUT_RATES = []
    CHIP_TYPE_NUM = 0
    
    class _Print(StringIO): 
        '''
        this is a inner class use for ui print to specify dut
        '''
        def __init__(self, factory_tool, obj):
            """ x.__init__(...) initializes x; see help(type(x)) for signature """
            self.factory_tool = factory_tool
            self.obj = obj
        
        def write(self, log):
            timestr=time.strftime('[%M:%S]',time.localtime(time.time()))
            log=timestr+log
            self.factory_tool.SIGNAL_PRINT.emit(self.obj, log)    
    
    def __init__(self, params={}, parent=None):
        super(QtGui.QMainWindow, self).__init__(parent=parent)
        self.setupUi(self)  # general by pyqt designer
        self.init_ui()     # add spicial control module and init some ui settings
        self.init_parameters()    # initial the settings for production test
        self.init_signal() # add signal for need
        
        
    ### init functions ###
    def init_ui(self):
        self.twTestArea.setCurrentIndex(0)
        for i in xrange(1,self.DUT_NUM+1):
            self.signal_print_log(eval('self.tbLog{}'.format(str(i))), '[state]idle')
            for j in xrange(1,3):
                self.DUT_PORTS.append(eval('self.cbPort'+str(i)+'_'+str(j)))
                self.DUT_RATES.append(eval('self.cbPortRate'+str(i)+'_'+str(j)))
                eval('self.lePortRate'+str(i)+'_'+str(j)).setHidden(True)
            

    def init_signal(self):
        QtCore.QObject.connect(self.trwTestFlow, QtCore.SIGNAL(_fromUtf8("itemChanged(QTreeWidgetItem*,int)")), self._signal_testflow_check)
        QtCore.QObject.connect(self.pbTFSubmit, QtCore.SIGNAL(_fromUtf8("clicked()")), self.button_testflow_submit)
        QtCore.QObject.connect(self.pbTFReset, QtCore.SIGNAL(_fromUtf8("clicked()")), self.button_testflow_reset)
        QtCore.QObject.connect(self.pbCloudSync, QtCore.SIGNAL(_fromUtf8("clicked()")), self.button_cloud_sync)
        QtCore.QObject.connect(self.pbDutReset, QtCore.SIGNAL(_fromUtf8("clicked()")), self.button_dut_reset)
        QtCore.QObject.connect(self.pbDutSubmit, QtCore.SIGNAL(_fromUtf8("clicked()")), self.button_dut_submit)
        QtCore.QObject.connect(self.cbChipType, QtCore.SIGNAL(_fromUtf8("currentIndexChanged(int)")), self.combobox_chip_type_change)
        QtCore.QObject.connect(self.cbTestFrom, QtCore.SIGNAL(_fromUtf8("currentIndexChanged(int)")), self.combobox_test_from_change)
        QtCore.QObject.connect(self.pbBinPath, QtCore.SIGNAL(_fromUtf8("clicked()")), self.button_showFileDialog)
        self.SIGNAL_PRINT.connect(self.signal_print_log)
        self.maCloud.changed.connect(self.button_change_position)        
        
        QtCore.QObject.connect(self.pbAllStart, QtCore.SIGNAL(_fromUtf8("clicked()")), self.button_all_start)
        QtCore.QObject.connect(self.pbAllStop, QtCore.SIGNAL(_fromUtf8("clicked()")), self.button_all_stop)        
        self.pbStart1.clicked.connect(lambda :self.button_single_start(self.pbStart1))
        self.pbStart2.clicked.connect(lambda :self.button_single_start(self.pbStart2))
        self.pbStart3.clicked.connect(lambda :self.button_single_start(self.pbStart3))
        self.pbStart4.clicked.connect(lambda :self.button_single_start(self.pbStart4))
        self.pbStop1.clicked.connect(lambda :self.button_single_stop(self.pbStop1))
        self.pbStop2.clicked.connect(lambda :self.button_single_stop(self.pbStop2))
        self.pbStop3.clicked.connect(lambda :self.button_single_stop(self.pbStop3))
        self.pbStop4.clicked.connect(lambda :self.button_single_stop(self.pbStop4))        
        
        for cb in self.DUT_PORTS: cb.clicked.connect(self.combobox_change_port) 
        self.cbPortRate1_1.currentIndexChanged.connect(lambda :self.combobox_change_baud(self.cbPortRate1_1))
        self.cbPortRate1_2.currentIndexChanged.connect(lambda :self.combobox_change_baud(self.cbPortRate1_2))
        self.cbPortRate2_1.currentIndexChanged.connect(lambda :self.combobox_change_baud(self.cbPortRate2_1))
        self.cbPortRate2_2.currentIndexChanged.connect(lambda :self.combobox_change_baud(self.cbPortRate2_2))
        self.cbPortRate3_1.currentIndexChanged.connect(lambda :self.combobox_change_baud(self.cbPortRate3_1))
        self.cbPortRate3_2.currentIndexChanged.connect(lambda :self.combobox_change_baud(self.cbPortRate3_2))
        self.cbPortRate4_1.currentIndexChanged.connect(lambda :self.combobox_change_baud(self.cbPortRate4_1))
        self.cbPortRate4_2.currentIndexChanged.connect(lambda :self.combobox_change_baud(self.cbPortRate4_2))
        
        self.maLog1.triggered.connect(lambda :self.button_pop_log(1))
        self.maLog2.triggered.connect(lambda :self.button_pop_log(2))
        self.maLog3.triggered.connect(lambda :self.button_pop_log(3))
        self.maLog4.triggered.connect(lambda :self.button_pop_log(4))
        
        QtCore.QMetaObject.connectSlotsByName(self)
    
    def init_parameters(self):
        self.esp_process={1:None, 2:None, 3:None, 4:None}
        self.run_flag = True
        self.mutex = threading.Lock()
        self.rfmutex = threading.Lock()
        
        self.dut_config = {}
        
        tmp_path = os.getcwd().replace('\\', '//')
        self.path_logs = tmp_path[:tmp_path.find(tmp_path.split('//')[len(tmp_path.split('//'))-1])] + 'logs//'
        self.path_threshold = tmp_path + '//config//'
        
        self.CHIP_TYPE_NUM = self.cbChipType.count()
        self.BAUD_NUM = self.cbPortRate1_1.count()
        self.button_dut_reset(file_path='./config/dutConfig', try_login=True) 
        
        self.test_flow = {}
        self.init_testflow()
        self.init_threshold()
        self.init_test_thread()
        
    def init_test_thread(self):
        for i in range(1,5):
            stdout_ = self._Print(self, eval("self.tbLog"+str(i)))
            self.dut_config['common_conf']['dut_num'] = str(i)
            self.esp_process[id]=esp_test.esp_testThread(stdout_, self.dut_config,self.test_flow, self.rfmutex)
    
    def init_threshold(self):
        import openpyxl
        self.dut_config['common_conf']['threshold_path'] = self.path_threshold
        
        if str(self.cbChipType.currentText()).find('32')>=0:
            wb = openpyxl.load_workbook(self.path_threshold+'full_Threshold_8266.xlsx')
        else:
            wb = openpyxl.load_workbook(self.path_threshold+'full_Threshold_8266.xlsx')

        sheet = wb.get_sheet_by_name(wb.sheetnames[0])
        
        for i in xrange(1, sheet.max_row+1):
            self.twThreshold.setRowCount(sheet.max_row)
            for j in xrange(1, sheet.max_column+1):
                twi = QtGui.QTableWidgetItem()
                if sheet[str(chr(64+j))+str(i)].value != None:
                    twi.setText(str(sheet[str(chr(64+j))+str(i)].value))
                    self.twThreshold.setItem(i-1, j-1, twi)
        
    def init_testflow(self):
        self.ui_update_testflow('./config/tmp_testFlow')
        if self.dut_config['common_conf']['position'] == 'cloud':
            flow_path = './config/cloudTestFlow'
        else:
            flow_path = './config/testFlow'
        try:            
            if self.button_testflow_reset(flow_path) != 0:
                print ('Test Flow file was broken, load with default config')
                self.button_testflow_reset('./config/tmp_testFlow')
        except:
            print ('Test Flow file was broken, load with default config')
            self.button_testflow_reset('./config/tmp_testFlow')
        self.ui_update_testflow(flow_path)
        os.remove('./config/tmp_testFlow')
        
    ### signal deal functions ###
    def _signal_testflow_check(self, item=None, index=None):
        def updateParentItem(item):
            try:
                parent = item.parent()
            except:
                return
            
            checkedCount = 0
            checkableCount = 0
            
            for i in xrange(parent.childCount()):
                childItem = parent.child(i)
                if childItem.flags()&QtCore.Qt.ItemIsUserCheckable:
                    checkableCount += 1
                    if childItem.checkState(0) == 2:
                        checkedCount += 1
            if(checkedCount <= 0):
                if parent.flags() & QtCore.Qt.ItemIsUserCheckable:
                    parent.setCheckState(0, QtCore.Qt.Unchecked)
            elif checkedCount>0 and checkedCount<checkableCount:
                if parent.flags() & QtCore.Qt.ItemIsUserCheckable:
                    parent.setCheckState(0, QtCore.Qt.PartiallyChecked)
            elif checkedCount == checkableCount:
                if parent.flags() & QtCore.Qt.ItemIsUserCheckable:
                    parent.setCheckState(0, QtCore.Qt.Checked)        
        
        if item.checkState(0) == QtCore.Qt.Unchecked:
            checkalbeChildCount = 0
            for i in xrange(item.childCount()):
                if item.child(i).flags() & QtCore.Qt.ItemIsUserCheckable:
                    item.child(i).setCheckState(0,QtCore.Qt.Unchecked)
                    checkalbeChildCount += 1
            if checkalbeChildCount <= 0:
                updateParentItem(item)
                
        elif item.checkState(0) == QtCore.Qt.Checked:
            checkalbeChildCount = 0
            for i in xrange(item.childCount()):
                if item.child(i).flags() & QtCore.Qt.ItemIsUserCheckable:
                    item.child(i).setCheckState(0,QtCore.Qt.Checked)
                    checkalbeChildCount += 1
            if checkalbeChildCount <= 0:
                updateParentItem(item)
                
    def signal_print_log(self, tb, log, err_code=0x0):
        show_flag = True
        log=str(log)
        state_flag = True
        dut_num = str(tb.objectName()[-1])
        if log.find('[state]') >= 0:
            show_flag = False
            if log.lower().find('idle') >= 0:
                state = 'IDLE'
                eval("self.tbLog"+str(dut_num)).clear()
                style = "background-color: rgb(0, 170, 255);\n"
            elif log.lower().find('sync') >= 0:
                state = 'SYNC'
                style = "background-color: rgb(0, 170, 255);\n"                
            elif log.lower().find('run') >= 0:
                state = 'RUN'
                style = "background-color: rgb(255, 255, 0);\n"
                
            elif log.lower().find('passed') >= 0:
                state = 'TESTED'
                style = "background-color: rgb(0, 170, 0);\n"
                if log.find('record') >= 0:
                    self._local_count('pass', dut_num)                
            elif log.lower().find('pass') >= 0:
                state = 'PASS'
                style = "background-color: rgb(0, 170, 0);\n"
                if log.find('record') >= 0:
                    self._local_count('pass', dut_num)
            elif log.lower().find('fail') >= 0:
                state = 'FAIL'
                style = "background-color: rgb(255, 0, 0);\n"
                if log.find('record') >= 0:
                    self._local_count('fail', dut_num, log.lower().split(',')[-1])
            elif log.lower().find('upload-f') >= 0:
                state = 'upload-f'
                style = "background-color: rgb(255, 0, 0);\n"
            elif log.lower().find('upload-p') >= 0:
                state = 'upload-p'
                style = "background-color: rgb(0, 0, 255);\n"    
            elif log.lower().find('finish') >= 0:
                state_flag = False
                eval('self.pbStart{}'.format(dut_num)).setDown(False)
                eval('self.pbStart{}'.format(dut_num)).setEnabled(True)
                    
            else:
                state_flag = False
                
            if state_flag:
                self._state_change(dut_num, state, style)
                                
        elif log.find('[upload]') >= 0:
            self.lbTotalStatus.setText(log.split(']')[-1])
        elif log.find('[mac]') >= 0:
            eval("self.lbMAC{}".format(dut_num)).setText(log[-12:])
        
        if show_flag:
            tb.append(log)
    
    def button_testflow_reset(self, file_path='./config/testFlow'):
        if self.dut_config['common_conf']['position'] == 'cloud':
            file_path = './config/cloudTestFlow'
        parent = 'root'
        with open(file_path, 'r') as fd:
            rl = fd.readline()
            while rl != '':
                if not rl.startswith('['):
                    rl = fd.readline()
                    continue
                rl = rl.strip().strip('[').strip(']')
                if len(rl.split(self._SP_SIGN)) < 5:
                    print('config file is broken, please re-generate')
                    return -1
                level_index, childCount, checkable, editable, value = rl.split(self._SP_SIGN)
                checkable = int(checkable.strip())
                editable = editable.strip()
                value = value.strip()
                level_index = level_index.strip().split('-')
                if editable == '1' or checkable >= 0:
                    item = self.trwTestFlow.invisibleRootItem()
                    for i in xrange(1,len(level_index)):
                        tmp = item.text(0)
                        if item.childCount() <= int(level_index[i]):
                            tmp_child = QtGui.QTreeWidgetItem()
                            tmp_child.setText(0,value)
                            item.addChild(tmp_child)
                            tmp_child.setHidden(True)
                        item = item.child(int(level_index[i]))
                    if editable == '1':
                        item.setText(0, value)
                        self.test_flow[parent] = value
                    if checkable >= 0:
                        item.setCheckState(0, checkable)
                        self.test_flow[value] = checkable
                
                parent = value
                rl = fd.readline()
                
        self.lbFWVer.setText(self.test_flow['USER_FW_VER_STR'])
        return 0
    
    def button_testflow_submit(self):
        first_flag = True
        while(1):
            tmp_time = time.strftime('%Y-%m-%d-%H',time.localtime(time.time()))
            y,m,d,h = map(lambda x:int(x), tmp_time.split('-'))
            print (tmp_time, (y+(m+d+h))%10000)
            
            if first_flag:
                verify,rst = QtGui.QInputDialog().getText(self, "Verify Box", "Verify:",
                                                          QtGui.QLineEdit().Normal, '')
            else:
                verify,rst = QtGui.QInputDialog().getText(self, "Verify Box", "Verify: (fail, retry!!)",
                                                          QtGui.QLineEdit().Normal, '')
            first_flag = False
            if rst:       
                try:
                    print(verify)
                    if int(verify) == (y+(m+d+h))%10000:
                        self.ui_update_testflow(pop_msg=True)
                        print ('verify pass')
                        break
                    else:
                        print ('verify fail')
                except:
                    print ('please input 4 bytes Number')
            else:
                break
    
    def button_all_start(self):
        for id in (1,2,3,4):
            try:
                if self.esp_process[id]==None or (not self.esp_process[id].isRunning()):
                    stdout_ = self._Print(self, eval("self.tbLog"+str(id)))
                    self.dut_config['common_conf']['dut_num'] = str(id)  
                    
                    self.esp_process[id]=esp_test.esp_testThread(stdout_, self.dut_config,self.test_flow, self.rfmutex)
                    self.esp_process[id].start()
                    eval('self.pbStart{}'.format(id)).setEnabled(False)
                    eval('self.pbStart{}'.format(id)).setDown(True)
                    time.sleep(0.1)
            except:
                pass
    
    def button_single_start(self, btn):
        btn.setEnabled(False)
        btn.setDown(True)
        id = int(btn.objectName()[len(btn.objectName())-1])
        if id in (1,2,3,4):
            stdout_ = self._Print(self, eval("self.tbLog"+str(id)))
            self.dut_config['common_conf']['dut_num'] = str(id)
            self.esp_process[id]=esp_test.esp_testThread(stdout_, self.dut_config,self.test_flow, self.rfmutex)
            self.esp_process[id].start()
        else:
            print('error: get strat btn err')
    
    def button_all_stop(self):
        for id in (1,2,3,4):
            try:
                self.esp_process[id].SIGNAL_STOP.emit()
                self.signal_print_log(eval('self.tbLog{}'.format(id)), '[state]idle')
                eval('self.lbMAC{}'.format(id)).setText('000000000000')
            except:
                self.signal_print_log(eval('self.tbLog{}'.format(id)), '[state]idle')
                eval('self.lbMAC{}'.format(id)).setText('000000000000')
    
    def button_single_stop(self, btn):
        id = int(btn.objectName()[len(btn.objectName())-1])
        eval('self.lbMAC{}'.format(id)).setText('000000000000')
        if self.esp_process.has_key(id):
            # if self.esp_process[id].isRunning():
            try:
                self.esp_process[id].SIGNAL_STOP.emit()
            except:
                pass
         
    def button_pop_log(self, index):
        log_path = self.path_logs
        if index in (1,2,3,4):       
            if self.esp_process[index] != None:
                log_path = self.esp_process[index].logpath.split('//')[-1]
                log_path = self.path_logs + log_path
            try:
                os.startfile(log_path)
            except:
                print('error path')
        else:
            print('error: get log btn err')
    
    def button_cloud_sync(self):
        self.lbSyncState.setText('try sync cloud config by mpn')
        
        self.pbCloudSync.setEnabled(False)
        self.pbCloudSync.setDown(True)
        conf = ConfigParser.ConfigParser()
        conf.read('./config/dutConfig')
        ip = conf.get('common_conf', 'tmp_server_ip')
        port = conf.get('common_conf', 'tmp_server_port')
        #url = 'http://{}:{}/hp_register.py?opration=config&user_token={}'.format(ip,port,self.leMPNNo.text()) # test_config_002
        url = 'https://{}:{}/mpn?mpnSid={}'.format(ip, port, str(self.leMPNNo.text())+'_'+str(self.dut_config["chip_conf"]['chip_type']))
           
        try:
            rsp = requests.get(url=url, verify=False,timeout=3).json()
            with open('./config/cloudTestFlow', 'w') as fd:
                fd.write(str(rsp['data']))
            #print str(rsp['data'])
        except:
            self.lbSyncState.setText('sync fail!!! please check the mpn and chip type')
            self.pbCloudSync.setEnabled(True)
            self.pbCloudSync.setDown(False)
            return
            
        try:
            self.button_testflow_reset(file_path='./config/cloudTestFlow')
            self.lbSyncState.setText('sync success')
        except:
            self.lbSyncState.setText('sync fail')
            self.button_testflow_reset(file_path='./config/testFlow')
        
        self.ui_update_dut()
        self.pbCloudSync.setEnabled(True)
        self.pbCloudSync.setDown(False)
        return

        
    def button_dut_reset(self, file_path='./config/dutConfig', try_login=False):
        conf = ConfigParser.ConfigParser()
        try:
            conf.read(file_path)
            # set dut_config
            for i in conf.sections():
                self.dut_config[i] = dict(conf.items(i))
            self.dut_config['common_conf']['threshold_path'] = self.path_threshold
                
            index = self.cbChipType.findText(self.dut_config['chip_conf']['chip_type'])
            if index >= 0:
                self.cbChipType.setCurrentIndex(index)  
                self.leChipType.setHidden(True)
            else:
                self.cbChipType.setCurrentIndex(self.CHIP_TYPE_NUM-1)
                self.leChipType.setText(self.dut_config['chip_conf']['chip_type'])
                self.leChipType.setHidden(False)
                
            self.lePoNo.setText(self.dut_config['common_conf']['po_no'])
            self.leMPNNo.setText(self.dut_config['common_conf']['mpn_no'])
            self.leFacId.setText(self.dut_config['common_conf']['fac_sid'])
            self.leBatchId.setText(self.dut_config['common_conf']['batch_sid'])
            self.leFacPlan.setText(self.dut_config['common_conf']['fac_plan'])
            self.leBinPath.setText(self.dut_config['common_conf']['bin_path'])
            
            if self.dut_config['common_conf']['test_from'] == 'RAM':
                self.cbTestFrom.setCurrentIndex(0)
            else:
                self.cbTestFrom.setCurrentIndex(1)
                
            self.cbFREQ.setCurrentIndex(self.cbFREQ.findText(self.dut_config['chip_conf']['freq']))
            self.cbAutoStart.setChecked(True if self.dut_config['common_conf']['auto_start'] == '1' else False)
            self.cbEfuseMode.setCurrentIndex(int(self.dut_config['chip_conf']['efuse_mode']))
                
            self.lbChipType.setText(self.dut_config['chip_conf']['chip_type'])
            self.lbPoNo.setText(self.lePoNo.text())
            self.lbMPNNo.setText(self.leMPNNo.text())
            self.lbBatchId.setText(self.leBatchId.text())
            self.lbFacPlan.setText(self.leFacPlan.text())
            self.lbAutoStart.setText('ON' if self.dut_config['common_conf']['auto_start'] == '1' else 'OFF')
            self.lbTestMode.setText(self.dut_config['common_conf']['test_from'])
            
            for i in xrange(1, self.DUT_NUM+1):
                for j in xrange(1, 3):
                    eval('self.cbPort'+str(i)+'_'+str(j)).setItemText(0, conf.get('DUT'+str(i), 'PORT'+str(j)))
                    eval('self.cbPort'+str(i)+'_'+str(j)).setCurrentIndex(0)
                    index = eval('self.cbPortRate'+str(i)+'_'+str(j)).findText(conf.get('DUT'+str(i), 'RATE'+str(j)))
                    if index >= 0:
                        eval('self.cbPortRate'+str(i)+'_'+str(j)).setCurrentIndex(index)
                        eval('self.lePortRate'+str(i)+'_'+str(j)).setHidden(True)
                    else:
                        eval('self.cbPortRate'+str(i)+'_'+str(j)).setCurrentIndex(self.BAUD_NUM-1)
                        eval('self.lePortRate'+str(i)+'_'+str(j)).setText(conf.get('DUT'+str(i), 'RATE'+str(j)))
                        eval('self.lePortRate'+str(i)+'_'+str(j)).setHidden(False)            
            
            if try_login:
                login_flag=False
                if self.dut_config['common_conf']['position'] == 'cloud':
                    token = self.dut_config['common_conf']['auth_token']
                    try:
                        if self._try_login(token)==0:
                            QtCore.QObject.disconnect(self.maCloud, QtCore.SIGNAL(_fromUtf8("changed()")), self.button_change_position)
                            self.maCloud.setChecked(True)
                            self.maCloud.changed.connect(self.button_change_position) 
                            self.wgCloudConfig.setHidden(False)
                            self.wgCloudShow.setHidden(False)
                            self.pbCloudSync.setEnabled(True)
                            #self.twTestArea.widget(2).setEnabled(False)
                            self.pbTFSubmit.setEnabled(False)
                            self.lbPosition.setText('Cloud')
                            self.tePosition.setStyleSheet(_fromUtf8("background-color: rgb(255, 255, 0);\n"
                                                                    "border-color: rgb(0, 255, 255);"))
                            login_flag=True
                        else:
                            print ('login fail')
                    except:
                        print ('login fail')
                if not login_flag:
                    self.maCloud.setChecked(False)
                    self.wgCloudConfig.setHidden(True)
                    self.wgCloudShow.setHidden(True)
                    self.pbCloudSync.setEnabled(False)
                    self.lbPosition.setText('Local')            
        except:
            print ('load to config file fail')
    
    def button_dut_submit(self):
        first_flag = True
        while(1):
            tmp_time = time.strftime('%Y-%m-%d-%H',time.localtime(time.time()))
            y,m,d,h = map(lambda x:int(x), tmp_time.split('-'))
            print (tmp_time, (y+(m+d+h))%10000)
            
            if first_flag:
                verify,rst = QtGui.QInputDialog().getText(self, "Verify Box", "Verify:",
                                                          QtGui.QLineEdit().Normal, '')
            else:
                verify,rst = QtGui.QInputDialog().getText(self, "Verify Box", "Verify: (fail, retry!!)",
                                                          QtGui.QLineEdit().Normal, '')
            first_flag = False
            if rst:       
                try:
                    print(verify)
                    if int(verify) == (y+(m+d+h))%10000:
                        print ('verify pass')
                        self.ui_update_dut()
                        self.init_threshold()
                        break
                    else:
                        print ('verify fail')
                except:
                    print ('please input 4 bytes Number')
            else:
                break
    
    def _try_login(self, token):
        ip = self.dut_config['common_conf']['tmp_server_ip']
        port = self.dut_config['common_conf']['tmp_server_port']
            
        url = 'https://{}:{}/factorys'.format(ip, port)
        headers = {'token':token}
        try:
            rsp = requests.get(url=url, headers=headers, verify=False, timeout=3).json()
            print(rsp)
            if rsp['status'] == 200:
                return 0
            else:
                print('login fail')
                return 1
        except:
            print('login fail')
            return 1
    
    def button_change_position(self):
        QtCore.QObject.disconnect(self.maCloud, QtCore.SIGNAL(_fromUtf8("changed()")), self.button_change_position)
        if self.maCloud.isChecked(): # local -> cloud
            first_flag = True
            while(1):
                if first_flag:
                    first_flag = False
                    token,rst = QtGui.QInputDialog().getText(self, "Login", "Token:",
                                                             QtGui.QLineEdit().Normal, '')
                else:
                    token,rst = QtGui.QInputDialog().getText(self, "Login", "Token: (fail, retry!!)",
                                                             QtGui.QLineEdit().Normal, '')
                if rst:
                    try:
                        if self._try_login(token)==0:
                            self.maCloud.setChecked(True)
                            self.maCloud.changed.connect(self.button_change_position)
                            self.wgCloudConfig.setHidden(False)
                            self.wgCloudShow.setHidden(False)
                            self.pbCloudSync.setEnabled(True)
                            #self.twTestArea.widget(2).setEnabled(False)
                            self.pbTFSubmit.setEnabled(False)
                            self.lbPosition.setText('Cloud')
                            self.tePosition.setStyleSheet(_fromUtf8("background-color: rgb(255, 255, 0);\n"
                                                                    "border-color: rgb(0, 255, 255);"))
                            conf = ConfigParser.ConfigParser()
                            conf.read('./config/dutConfig')
                            conf.set('common_conf', 'auth_token', str(token))
                            conf.write(open('./config/dutConfig', 'w'))
                            self.ui_update_dut()
                            self.button_testflow_reset(file_path='./config/cloudTestFlow')
                            return
                        else:
                            print ('login fail')
                    except:
                        print ('login fail')
                else:
                    break
                
        self.maCloud.setChecked(False)
        self.wgCloudConfig.setHidden(True)
        self.wgCloudShow.setHidden(True)
        self.maCloud.changed.connect(self.button_change_position) 
        self.pbCloudSync.setEnabled(False)
        #self.twTestArea.widget(2).setEnabled(True)
        self.pbTFSubmit.setEnabled(True)
        self.lbPosition.setText('Local')
        self.tePosition.setStyleSheet(_fromUtf8("background-color: rgb(255, 170, 127);\n"
                                                "border-color: rgb(0, 255, 255);"))
        self.ui_update_dut()
        self.button_testflow_reset()
    
    def button_showFileDialog(self):
        filename = QtGui.QFileDialog.getOpenFileName(None, 'Open file', './bin/', filter='firmware(*.bin);;all(*.*)', selectedFilter='firmware(*.bin)')
        self.leBinPath.setText(filename)
        self.dut_config['common_conf']['bin_path'] = filename
        
    def combobox_change_port(self, cb_port):
        port_list = list(serial.tools.list_ports.comports())
        cb_port.clear()
        for port in port_list:
            print(port[0])
            cb_port.addItem(_fromUtf8(port[0]))
        cb_port.showPopup()
    
    def combobox_change_baud(self, cb):
        if cb.currentText() == 'custom':
            eval('self.lePortRate{}'.format(cb.objectName()[-3:])).setHidden(False)
        else:
            eval('self.lePortRate{}'.format(cb.objectName()[-3:])).setHidden(True)
     
    def combobox_chip_type_change(self, index):
        ''' @new remove input module names operation
        '''
        if True:
            return
        if index == self.CHIP_TYPE_NUM - 1:
            self.leChipType.setHidden(False)
        else:
            self.leChipType.setHidden(True)
    
    def combobox_test_from_change(self, index):
        self.dut_config['common_conf']['test_from'] = self.cbTestFrom.currentText()
        if self.cbTestFrom.currentText() == 'RAM':
            self.pbBinPath.setEnabled(True)
            self.leBinPath.setEnabled(True)
        else:
            self.pbBinPath.setEnabled(False)
            self.leBinPath.setEnabled(False)            
    
    ### class functions ###
    def ui_update_testflow(self, file_path='./config/testFlow', pop_msg=False):
        with open(file_path, 'w') as fd:
            fd.write("- level-index $$ childCount$$ checkable$$ editable$$ value -\n")
            self._testflow_general(fd, self.trwTestFlow.invisibleRootItem(), '0')
        self.lbFWVer.setText(self.test_flow['USER_FW_VER_STR'])
        if pop_msg:
            msg = QtGui.QMessageBox(QtGui.QMessageBox.NoIcon, '!!!','The Test Flow update succ!!    ')
            msg.exec_()
    
    def _testflow_general(self, fd, root, level_index):
        if(root.flags()&QtCore.Qt.ItemIsEditable):            
            self.test_flow[str(root.parent().text(0))] = str(root.text(0))
        else:
            self.test_flow[str(root.text(0))] = root.checkState(0)
            
        fd.write("[%-12s%s%2d%s%2d%s%2d%s %s]\n"%(level_index, self._SP_SIGN, int(root.childCount()), self._SP_SIGN, 
                                                  int(root.checkState(0) if(root.flags()&QtCore.Qt.ItemIsUserCheckable) else -1), self._SP_SIGN, 
                                                  int(1 if(root.flags()&QtCore.Qt.ItemIsEditable) else 0), self._SP_SIGN, str(root.text(0))))
        for i in xrange(root.childCount()):
            self._testflow_general(fd, root.child(i), level_index+'-'+str(i))
    
    def ui_update_dut(self, file_path='./config/dutConfig'):
        conf = ConfigParser.ConfigParser()
        conf.read(file_path)
        if not conf.has_section('common_conf'):
            conf.add_section('common_conf')
        if not conf.has_section('chip_conf'):
            conf.add_section('chip_conf')
            
        conf.set('common_conf', 'POSITION', 'cloud' if self.maCloud.isChecked() else 'local')
        conf.set('common_conf', 'TEST_FROM', self.cbTestFrom.currentText())
        conf.set('chip_conf', 'FREQ', self.cbFREQ.currentText())
        conf.set('chip_conf', 'efuse_mode', self.cbEfuseMode.currentIndex())
        conf.set('common_conf', 'auto_start', '1' if self.cbAutoStart.checkState() else '0')
        if self.cbChipType.currentIndex() < self.CHIP_TYPE_NUM:
            conf.set('chip_conf', 'CHIP_TYPE', self.cbChipType.currentText())
        else:
            conf.set('chip_conf', 'CHIP_TYPE', self.leChipType.text())
        conf.set('common_conf', 'PO_NO', self.lePoNo.text())
        conf.set('common_conf', 'MPN_NO', self.leMPNNo.text())
        conf.set('common_conf', 'FAC_SID', self.leFacId.text())
        conf.set('common_conf', 'BATCH_SID', self.leBatchId.text())
        conf.set('common_conf', 'FAC_PlAN', self.leFacPlan.text())
        conf.set('common_conf', 'BIN_PATH', self.leBinPath.text())
        
        for i in xrange(1, self.DUT_NUM+1):
            if not conf.has_section('DUT'+str(i)):
                conf.add_section('DUT'+str(i))
            for j in xrange(1, 3):
                conf.set('DUT'+str(i), 'PORT'+str(j), str(eval('self.cbPort'+str(i)+'_'+str(j)).currentText()))
                if eval('self.cbPortRate'+str(i)+'_'+str(j)).currentIndex() < self.BAUD_NUM-1:
                    conf.set('DUT'+str(i), 'RATE'+str(j), str(eval('self.cbPortRate'+str(i)+'_'+str(j)).currentText()))
                else:
                    conf.set('DUT'+str(i), 'RATE'+str(j), str(eval('self.lePortRate'+str(i)+'_'+str(j)).text()))
        conf.write(open(file_path, 'w'))
        
        self.button_dut_reset(file_path)
        #conf.read(file_path)
        #for i in conf.sections():
            #self.dut_config[i] = dict(conf.items(i))
            
        if file_path=='./config/dutConfig':
            msg = QtGui.QMessageBox(QtGui.QMessageBox.NoIcon, '!!!','The DUT config update succ!!    ')
            msg.exec_()    
    
    def _state_change(self, dut_num, state, style):
        leState = eval("self.leStatus{}".format(dut_num))
        leState.setText(state)
        leState.setStyleSheet(_fromUtf8(style+"color: rgb(255, 255, 255);"))
        
    def _local_count(self, rst, dut_num, err_code=0x0):
        self.mutex.acquire()
        # total, pass, fail, mac, time
        datas = [0,0,0,0,0,0]
        mac = eval("self.lbMAC{}".format(dut_num)).text()
        count_path = './config/temp/'+self.dut_config['common_conf']['fac_plan']+'_count.txt'
        try:
            if not os.path.exists(count_path):
                try:
                    f=open(count_path, 'w')
                    f.close()
                except:
                    print('create count file fail')
                    return
                    
            with open(count_path, 'a+') as fd:
                rls = fd.readlines()
                
                if len(rls) > 0:
                    for i in xrange(1, len(rls)+1 if len(rls)<3 else 4):
                        if len(rls[-1 * i].split(',')) >= 4:
                            datas = rls[-1].split(',')
                            break
                        
                try:
                    total = int(datas[0])
                    pass_num = int(datas[1])
                    fail_num = int(datas[2])
                except:
                    total = 0
                    pass_num = 0
                    fail_num = 0
                
                now_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')[5:]
                if rst == 'pass':
                    pass_num += 1
                elif rst == 'fail':
                    fail_num += 1
                total += 1
                
                fd.write(str(total)+','+str(pass_num)+','+str(fail_num)+','+str(mac)+','+now_time+','+str(err_code)+"\n")
        except:
            print('open count file fail, please release it')
        
        try:   
            with open('./mac_list/'+self.dut_config['common_conf']['fac_plan']+'_'+rst+'.csv', 'a') as fd:
                fd.write(mac+','+str(err_code)+'\n')
        except:
            print('open mac_list fail, please release it')
        
        self.mutex.release()
        self.lbWorkStat.setText('pass:{}/ fail:{}'.format(pass_num, fail_num))
              

def main():
    app = QtGui.QApplication(sys.argv)
    ui = FactoryToolUI()
    ui.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()